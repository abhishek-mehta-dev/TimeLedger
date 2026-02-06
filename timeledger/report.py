"""
Report module for TimeLedger - CSV report generation.
Generates daily reports with summary and event timeline.
"""

import os
import hashlib
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .db import get_events_for_date, store_report_hash
from .tracker import WorkTracker


def format_duration(seconds: float) -> str:
    """Format seconds as HH:MM:SS."""
    if seconds < 0:
        seconds = 0
    
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = int(seconds % 60)
    
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


def format_time(timestamp_str: str) -> str:
    """Format ISO timestamp to local readable time."""
    try:
        dt = datetime.fromisoformat(timestamp_str.replace("Z", "+00:00"))
        local_dt = dt.astimezone()
        return local_dt.strftime("%I:%M:%S %p")
    except (ValueError, AttributeError, TypeError):
        return timestamp_str if timestamp_str else "N/A"


def generate_report(
    date: str,
    output_dir: Optional[str] = None,
    store_hash: bool = True
) -> str:
    """
    Generate a high-end styled Excel report for the specified date.
    """
    # Get events and stats
    events = get_events_for_date(date)
    tracker = WorkTracker()
    stats = tracker.get_stats_for_date(date)
    
    # Pre-calculate productivity ratio
    productivity_percent = 0.0
    if stats.total_span_seconds > 0:
        productivity_percent = (stats.work_seconds / stats.total_span_seconds) * 100

    # Format date for header
    try:
        dt_header = datetime.strptime(date, "%Y-%m-%d")
        display_date = dt_header.strftime("%A, %B %d, %Y")
    except:
        display_date = date

    # Prepare output path
    if output_dir is None:
        output_dir = os.getcwd()
    
    filename = f"{date}-TimeLedger.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary Dashboard"
    
    # --- Define Colors (Consistent with UI) ---
    UI_PRIMARY = "3B82F6"   # Strong Blue
    UI_SUCCESS = "10B981"   # Emerald Green
    UI_DANGER = "EF4444"    # Rose Red
    UI_TEXT = "1E293B"      # Slate Dark
    UI_MUTED = "64748B"     # Slate Gray
    UI_BG_DARK = "1E293B"   # Header Background
    UI_STRIPE = "F8FAFC"    # Zebra Stripe
    
    # --- Styles ---
    title_font = Font(name='Segoe UI', size=20, bold=True, color="FFFFFF")
    header_font = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")
    label_font = Font(name='Segoe UI', size=10, bold=True, color=UI_TEXT)
    normal_font = Font(name='Segoe UI', size=10, color=UI_TEXT)
    footer_font = Font(name='Segoe UI', italic=True, size=9, color=UI_MUTED)
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', indent=1)
    
    fill_header = PatternFill(start_color=UI_BG_DARK, end_color=UI_BG_DARK, fill_type="solid")
    fill_section = PatternFill(start_color=UI_PRIMARY, end_color=UI_PRIMARY, fill_type="solid")
    fill_stripe = PatternFill(start_color=UI_STRIPE, end_color=UI_STRIPE, fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color="CBD5E1"),
        right=Side(style='thin', color="CBD5E1"),
        top=Side(style='thin', color="CBD5E1"),
        bottom=Side(style='thin', color="CBD5E1")
    )

    # --- Header ---
    ws.merge_cells('A1:D1')
    cc = ws['A1']
    cc.value = "TIMELEDGER PERFORMANCE DASHBOARD"
    cc.font = title_font
    cc.fill = fill_header
    cc.alignment = center_align
    ws.row_dimensions[1].height = 50
    
    ws.merge_cells('A2:D2')
    cc2 = ws['A2']
    cc2.value = display_date
    cc2.font = Font(name='Segoe UI', size=12, italic=True, color="FFFFFF")
    cc2.fill = fill_header
    cc2.alignment = center_align
    ws.row_dimensions[2].height = 25
    
    # --- Productivity Snapshot ---
    curr_row = 4
    ws.cell(row=curr_row, column=1, value="PRODUCTIVITY SNAPSHOT").font = header_font
    ws.cell(row=curr_row, column=1).fill = fill_section
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
    curr_row += 1
    
    summary_data = [
        ("Total Work Time", format_duration(stats.work_seconds)),
        ("Total Break Time", format_duration(stats.break_seconds)),
        ("Number of Breaks", f"{stats.break_count} sessions"),
        ("Productivity Ratio", f"{productivity_percent:.1f}%"),
    ]
    
    for key, value in summary_data:
        c1 = ws.cell(row=curr_row, column=1, value=key)
        c2 = ws.cell(row=curr_row, column=2, value=value)
        c1.font = label_font
        c2.font = Font(name='Segoe UI', size=10, bold=True, color=UI_PRIMARY if "Ratio" in key else UI_TEXT)
        c1.border = border_thin
        c2.border = border_thin
        c1.alignment = left_align
        c2.alignment = center_align
        curr_row += 1
        
    # --- Session Windows ---
    curr_row += 1
    ws.cell(row=curr_row, column=1, value="SESSION WINDOWS").font = header_font
    ws.cell(row=curr_row, column=1).fill = fill_section
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
    curr_row += 1
    
    session_data = [
        ("First Activity", format_time(stats.first_start.isoformat()) if stats.first_start else "N/A"),
        ("Final Activity", format_time(stats.last_end.isoformat()) if stats.last_end else "N/A"),
        ("Total Daily Span", format_duration(stats.total_span_seconds)),
    ]
    
    for key, value in session_data:
        c1 = ws.cell(row=curr_row, column=1, value=key)
        c2 = ws.cell(row=curr_row, column=2, value=value)
        c1.font = label_font
        c2.font = normal_font
        c1.border = border_thin
        c2.border = border_thin
        c1.alignment = left_align
        c2.alignment = center_align
        curr_row += 1

    # --- Break Explanations ---
    if stats.break_reasons:
        curr_row += 1
        ws.cell(row=curr_row, column=1, value="BREAK RATIONALE").font = header_font
        ws.cell(row=curr_row, column=1).fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
        curr_row += 1
        for i, reason in enumerate(stats.break_reasons, 1):
            ws.cell(row=curr_row, column=1, value=f"Pause {i}").font = label_font
            ws.cell(row=curr_row, column=2, value=reason).font = normal_font
            ws.cell(row=curr_row, column=1).border = border_thin
            ws.cell(row=curr_row, column=2).border = border_thin
            curr_row += 1

    # --- Timeline ---
    curr_row += 1
    ws.cell(row=curr_row, column=1, value="ACTIVITY LOG & TIMELINE").font = header_font
    ws.cell(row=curr_row, column=1).fill = fill_section
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=4)
    curr_row += 1
    
    headers = ["ID", "CLOCK TIME", "ACTION PERFORMANCE", "LOG DETAILS / REASON"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=curr_row, column=col, value=h)
        c.font = label_font
        c.border = border_thin
        c.alignment = center_align
        c.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        
    curr_row += 1
    for i, event in enumerate(events, 1):
        is_stripe = (i % 2 == 0)
        row_cells = []
        row_cells.append(ws.cell(row=curr_row, column=1, value=i))
        row_cells.append(ws.cell(row=curr_row, column=2, value=format_time(event.get("timestamp", ""))))
        
        action = event.get("action", "")
        action_cell = ws.cell(row=curr_row, column=3, value=action)
        row_cells.append(action_cell)
        
        # Color specific actions
        if action == "START": action_cell.font = Font(name='Segoe UI', bold=True, color=UI_SUCCESS)
        elif action == "END": action_cell.font = Font(name='Segoe UI', bold=True, color=UI_DANGER)
        elif action == "PAUSE": action_cell.font = Font(name='Segoe UI', bold=True, color="F59E0B")
        
        row_cells.append(ws.cell(row=curr_row, column=4, value=event.get("reason", "-")))
        
        for c in row_cells:
            c.border = border_thin
            if not c.font.bold: c.font = normal_font
            c.alignment = center_align if c.column < 4 else left_align
            if is_stripe: c.fill = fill_stripe
            
        curr_row += 1
        
    # --- Footer ---
    curr_row += 2
    
    # Use a hash of the event data for verification instead of the file itself to avoid circular dependency
    event_summary = "|".join([f"{e.get('timestamp')}:{e.get('action')}" for e in events])
    data_hash = hashlib.sha256(event_summary.encode()).hexdigest()
    
    f_cell = ws.cell(row=curr_row, column=1, value=f"Verification Hash: {data_hash[:12]}... (Data Integrity Check)")
    f_cell.font = footer_font
    curr_row += 1
    f2_cell = ws.cell(row=curr_row, column=1, value=f"Made by Abhishek Mehta | Exported from TimeLedger v2.0 on {datetime.now().strftime('%Y-%m-%d at %I:%M %p')}")
    f2_cell.font = footer_font

    # --- Auto-adjust columns ---
    for col in ws.columns:
        max_length = 0
        column_idx = col[0].column
        column_letter = get_column_letter(column_idx)
        for cell in col:
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column_letter].width = min(max_length + 4, 50)
        
    wb.save(filepath)
    
    # Calculate and store hash (using file hash for DB storage is fine now that file is saved)

    
    # Calculate and store hash
    if store_hash:
        try:
            sha256_hash = calculate_file_hash(filepath)
            store_report_hash(date, filename, sha256_hash)
        except Exception:
            pass
            
    return filepath


def calculate_file_hash(filepath: str) -> str:
    """Calculate SHA256 hash of a file."""
    sha256 = hashlib.sha256()
    with open(filepath, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            sha256.update(chunk)
    return sha256.hexdigest()


def generate_today_report(output_dir: Optional[str] = None) -> str:
    """Generate report for today."""
    today = datetime.now().strftime("%Y-%m-%d")
    return generate_report(today, output_dir)
